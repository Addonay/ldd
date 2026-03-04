from __future__ import annotations

from pathlib import Path
from time import time
from typing import Any

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================================
# CONFIGURATION CONSTANTS
# ============================================================================

# Input Excel file path
INPUT_FILE = r"C:\Users\23225959\dev\ldd\static\reports\KE_QoS_Report_20260212.xlsx"

# Output directory and file pattern
OUTPUT_DIR = r"C:\Users\23225959\dev\ldd\static\reports\generated"
OUTPUT_FILENAME_PREFIX = "KE_QoS_Report_20260212_charts"

# Sheet name to read from
SHEET_NAME = "KE"

# Number of recent date columns to analyze
RECENT_COLS = 30

# Hard limit: only inspect first 403 data rows
DATA_ROW_LIMIT = 403

# Fixed row slices provided by user (0-based, end exclusive)
REGION_SLICES: dict[str, tuple[int, int]] = {
    "kenya": (0, 57),
    "nairobi": (59, 125),
    "mombasa": (127, 193),
    "nakuru": (195, 261),
    "kisumu": (263, 329),
    "eldoret": (331, 397),
}

# Chart layout constants
CHART_START_COLS = [1, 10]
CHARTS_PER_ROW = 2
ROW_HEIGHT_BLOCK = 16


# Default thresholds (extracted from default-thresholds.ts)
DEFAULT_THRESHOLDS: dict[str, tuple[str, float]] = {
    # RNA/Accessibility
    '2G RNA': ('>', 99.93),
    '3G RNA': ('>', 99.93),
    '4G RNA': ('>', 99.93),
    '5G RNA': ('>', 99.93),
    
    # Drop rates
    '2G DCR': ('<=', 0.3),
    '3G CS DROP': ('<=', 0.15),
    '3G PS/HS DROP': ('<=', 0.3),
    'Service Drop Rate (All) (4G)': ('<=', 0.10),
    'VoLTE Drop Rate (QC1)': ('<=', 0.13),
    '5G Session Drop Rate': ('<=', 0.15),
    
    # Utilization
    '2G Utilization': ('<=', 80),
    '3G Utilization': ('<=', 80),
    '4G Utilization': ('<=', 80),
    '5G Utilization': ('<=', 80),
    
    # CSSR
    '2G CSSR': ('>=', 99.5),
    '3G CSSR': ('>=', 99.5),
    '3G HS/PS CSSR': ('>=', 99.0),
    'VoLTE CSSR': ('>=', 99.5),
    
    # PSR
    '2G PSR': ('>=', 90.0),
    '3G PSR': ('>=', 90.0),
    '4G PSR': ('>=', 90.0),
    
    # Success Rates
    'CSFB success rate (4G)': ('>=', 99.90),
    'Session Success Rate (4G)': ('>=', 99.5),
    'ERAB Setup Success Rate (QCI 1) -VoLTE': ('>=', 99.5),
    'SRVCC Success Rate (3G & 2G)': ('>=', 97.5),
    'SgNb Addition Success Rate': ('>=', 99.5),
    
    # Throughput
    '3G Throughput': ('>=', 2500),
    '4G Throughput(DL_User_throughput) (Mbps)': ('>=', 10),
    '5G DL Throughput (Mbps)': ('>=', 20),
    '5G UL Throughput (Mbps)': ('>=', 5),
    
    # CQI
    '4G CQI': ('>=', 8),
    '5G CQI': ('>=', 8),
    
    # TCH
    '2G_TCH_Availability': ('>=', 99.5),
    '2G_TCH_Blocking': ('<=', 0.15),
    
    # Data Integrity
    '2G voice Data integrity(%)': ('>=', 99.90),
    '3G CS Data integrity(%)': ('>=', 99.90),
    '3G PS Data integrity(%)': ('>=', 99.90),
    '4G Data integrity(%)': ('>=', 99.90),
    '5G Data integrity(%)': ('>=', 99.90),
    
    # Cell Performance
    '% of cells having 3G user throughput greater than 1000kbps': ('>=', 85.00),
    '% of cells having 4G user throughput greater than 5Mbps': ('>=', 90.00),
    '% of cells having 5G user throughput (>=20Mbps)': ('>=', 70.00),
    '% of cells with 2G RX Qual Samples 0-5(>=90%)': ('>=', 90.00),
    
    # Latency & Loss
    '5G Latency': ('<', 30),
    '5G Packet Loss': ('<', 0.01),
    
    # Other
    'Percentage_Cells_with_4G_User_Throughput_less_than_3Mbps': ('<', 5),
    'SD Blocking': ('<', 0.30)
}


def meets_threshold(operator: str, actual: float, target: float) -> bool:
    """Check if actual value meets the threshold requirement."""
    if operator == ">=":
        return actual >= target
    if operator == "<=":
        return actual <= target
    if operator == ">":
        return actual > target
    if operator == "<":
        return actual < target
    if operator == "==":
        return abs(actual - target) < 1e-9
    if operator == "!=":
        return abs(actual - target) >= 1e-9
    return True


def to_float(value: Any) -> float:
    """Convert value to float safely (NaN/invalid -> 0)."""
    if value is None:
        return 0.0
    if isinstance(value, float) and value != value:
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def format_threshold_value(value: float) -> str:
    """Format threshold value for display."""
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def calc_y_bounds(values: list[float], threshold_value: float) -> tuple[float, float]:
    """Calculate y-axis bounds with padding."""
    all_values = values + [threshold_value]
    min_v = min(all_values)
    max_v = max(all_values)
    span = max_v - min_v
    padding = (span * 0.15) if span > 0 else max(abs(max_v) * 0.1, 1.0)
    return min_v - padding, max_v + padding


def read_and_process_data(excel_path: Path, sheet_name: str, num_recent_cols: int) -> tuple[dict[str, dict[str, tuple[int, list[float]]]], list[str]]:
    """
    Read KE sheet with openpyxl (capped to first DATA_ROW_LIMIT rows), then map
    rows into fixed region slices and extract recent KPI values.
    """
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    ws = workbook[sheet_name]

    rows = list(ws.iter_rows(min_row=1, max_row=DATA_ROW_LIMIT + 1, values_only=True))
    workbook.close()

    if not rows or len(rows[0]) < 2:
        return {}, []

    headers = list(rows[0])
    date_col_indexes = list(range(max(len(headers) - num_recent_cols, 0), len(headers)))
    date_cols = [str(headers[i]) for i in date_col_indexes]
    kpi_col_idx = 1

    result: dict[str, dict[str, tuple[int, list[float]]]] = {region: {} for region in REGION_SLICES}

    for region, (start, end) in REGION_SLICES.items():
        for rel_idx, data_idx in enumerate(range(start, end)):
            excel_row_idx = data_idx + 2
            row_list_idx = excel_row_idx - 1
            if row_list_idx >= len(rows):
                break

            row = rows[row_list_idx]
            kpi_name = str((row[kpi_col_idx] if kpi_col_idx < len(row) else "") or "").strip()
            if not kpi_name:
                continue

            values = [to_float(row[i] if i < len(row) else None) for i in date_col_indexes]
            result[region][kpi_name] = (excel_row_idx, values)

    return result, date_cols


def generate_charts() -> dict[str, Any]:
    """Generate charts workbook from Excel data."""

    timestamp = int(time() * 1000)  # milliseconds since epoch
    output_filename = f"{OUTPUT_FILENAME_PREFIX}_{timestamp}.xlsx"
    output_path = Path(OUTPUT_DIR) / output_filename

    print(f"Reading data from {INPUT_FILE}, sheet '{SHEET_NAME}'...")
    region_data, date_cols = read_and_process_data(Path(INPUT_FILE), SHEET_NAME, RECENT_COLS)

    print(f"Using {len(region_data)} fixed regions and {len(date_cols)} recent date columns")

    workbook = load_workbook(INPUT_FILE)
    data_ws = workbook[SHEET_NAME]

    charts_sheet_name = "Charts"
    if charts_sheet_name in workbook.sheetnames:
        workbook.remove(workbook[charts_sheet_name])

    charts_ws = workbook.create_sheet(charts_sheet_name, 0)

    for col in range(1, 20):
        charts_ws.column_dimensions[get_column_letter(col)].width = 11

    chart_count = 0
    region_count = 0
    row_cursor = 1

    all_cols_in_sheet = [cell.value for cell in data_ws[1]]
    date_col_indices = []
    for date_col_name in date_cols:
        try:
            idx = all_cols_in_sheet.index(date_col_name) + 1
            date_col_indices.append(idx)
        except ValueError:
            continue

    if not date_col_indices:
        print("Error: Could not find date columns in sheet")
        return {"output_path": str(output_path), "regions_with_failures": 0, "charts_created": 0}

    min_date_col = min(date_col_indices)
    max_date_col = max(date_col_indices)

    for region_name, kpis in region_data.items():
        failing_kpis: list[tuple[str, str, float, int, list[float]]] = []

        for kpi_name, (row_idx, values) in kpis.items():
            if kpi_name not in DEFAULT_THRESHOLDS:
                continue

            operator, threshold_value = DEFAULT_THRESHOLDS[kpi_name]
            if not values:
                continue

            if not meets_threshold(operator, values[-1], threshold_value):
                failing_kpis.append((kpi_name, operator, threshold_value, row_idx, values))

        if not failing_kpis:
            continue

        print(f"  {region_name}: {len(failing_kpis)} failing KPIs")
        region_count += 1

        charts_ws.merge_cells(f"A{row_cursor}:R{row_cursor}")
        title_cell = charts_ws[f"A{row_cursor}"]
        title_cell.value = region_name.upper()
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        grid_start_row = row_cursor + 2

        for index, (kpi_name, operator, threshold_value, kpi_row_idx, values) in enumerate(failing_kpis):
            chart = LineChart()
            chart.title = f"{kpi_name} ({operator} {format_threshold_value(threshold_value)})"
            chart.style = 2
            chart.height = 8
            chart.width = 17
            chart.legend = None
            chart.x_axis.number_format = "mmm d"
            chart.x_axis.title = None
            chart.y_axis.title = None

            y_min, y_max = calc_y_bounds(values, threshold_value)
            chart.y_axis.scaling.min = y_min
            chart.y_axis.scaling.max = y_max

            values_ref = Reference(
                data_ws,
                min_col=min_date_col,
                max_col=max_date_col,
                min_row=kpi_row_idx,
                max_row=kpi_row_idx
            )
            categories_ref = Reference(
                data_ws,
                min_col=min_date_col,
                max_col=max_date_col,
                min_row=1,
                max_row=1
            )

            chart.add_data(values_ref, titles_from_data=False, from_rows=True)
            chart.set_categories(categories_ref)

            if chart.series:
                series = chart.series[0]
                series.smooth = False
                series.marker.symbol = "circle"
                series.marker.size = 4
                series.graphicalProperties.line.solidFill = "2E75B6"
                series.graphicalProperties.line.width = 24000

            grid_row = index // CHARTS_PER_ROW
            grid_col = index % CHARTS_PER_ROW
            anchor_row = grid_start_row + (grid_row * ROW_HEIGHT_BLOCK)
            anchor_col = CHART_START_COLS[grid_col]
            anchor_cell = f"{get_column_letter(anchor_col)}{anchor_row}"

            charts_ws.add_chart(chart, anchor_cell)
            chart_count += 1

        rows_used = ((len(failing_kpis) - 1) // CHARTS_PER_ROW + 1) * ROW_HEIGHT_BLOCK
        row_cursor = grid_start_row + rows_used + 2

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    print(f"\nGenerated {chart_count} charts across {region_count} regions")
    print(f"Output saved to: {output_path}")

    return {
        "output_path": str(output_path),
        "regions_with_failures": region_count,
        "charts_created": chart_count,
    }


def main() -> None:
    """Main entry point."""
    input_path = Path(INPUT_FILE)
    
    if not input_path.exists():
        print(f"Error: Input file not found: {input_path}")
        return
    
    result = generate_charts()
    
    print(f"\n✓ Success!")
    print(f"  Output: {result['output_path']}")
    print(f"  Regions: {result['regions_with_failures']}")
    print(f"  Charts: {result['charts_created']}")


if __name__ == "__main__":
    main()

