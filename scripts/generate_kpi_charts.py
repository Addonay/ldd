from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


@dataclass
class Threshold:
	operator: str
	value: float


def _to_datetime(value: Any) -> datetime | None:
	if isinstance(value, datetime):
		return value
	if isinstance(value, str):
		text = value.strip()
		if not text:
			return None
		text = text.replace("Z", "+00:00")
		try:
			parsed = datetime.fromisoformat(text)
			if parsed.tzinfo is not None:
				return parsed.replace(tzinfo=None)
			return parsed
		except ValueError:
			return None
	return None


def _to_float(value: Any) -> float | None:
	if value is None:
		return None
	try:
		return float(value)
	except (TypeError, ValueError):
		return None


def _meets(operator: str, actual: float, target: float) -> bool:
	if operator == ">=":
		return actual >= target
	if operator == "<=":
		return actual <= target
	if operator == ">":
		return actual > target
	if operator == "<":
		return actual < target
	if operator == "==":
		return abs(actual - target) < 1e-9
	if operator == "!=":
		return abs(actual - target) >= 1e-9
	return True


def _latest_numeric_value(points: list[dict[str, Any]]) -> float | None:
	best_date: datetime | None = None
	best_value: float | None = None
	for point in points:
		date = _to_datetime(point.get("date"))
		value = _to_float(point.get("value"))
		if date is None or value is None:
			continue
		if best_date is None or date > best_date:
			best_date = date
			best_value = value
	return best_value


def _sorted_points(points: list[dict[str, Any]]) -> list[tuple[datetime, float]]:
	parsed: list[tuple[datetime, float]] = []
	for point in points:
		date = _to_datetime(point.get("date"))
		value = _to_float(point.get("value"))
		if date is None or value is None:
			continue
		parsed.append((date, value))
	parsed.sort(key=lambda item: item[0])
	return parsed


def _build_thresholds(raw: dict[str, Any]) -> dict[str, Threshold]:
	out: dict[str, Threshold] = {}
	for name, item in raw.items():
		if not isinstance(item, dict):
			continue
		op = str(item.get("operator", "")).strip()
		val = _to_float(item.get("value"))
		if op and val is not None:
			out[name] = Threshold(operator=op, value=val)
	return out


def generate(payload: dict[str, Any], output_path: Path, source_path: Path | None = None) -> dict[str, Any]:
	thresholds = _build_thresholds(payload.get("thresholds", {}))
	areas: list[dict[str, Any]] = payload.get("areas", [])

	if source_path and source_path.exists():
		workbook = load_workbook(source_path)
	else:
		workbook = Workbook()

	charts_sheet_name = "Charts"
	data_sheet_name = "_ChartData"

	if charts_sheet_name in workbook.sheetnames:
		workbook.remove(workbook[charts_sheet_name])
	if data_sheet_name in workbook.sheetnames:
		workbook.remove(workbook[data_sheet_name])

	if len(workbook.sheetnames) == 1 and workbook.sheetnames[0] == "Sheet":
		workbook.remove(workbook["Sheet"])

	charts_ws = workbook.create_sheet(charts_sheet_name)
	data_ws = workbook.create_sheet(data_sheet_name)
	data_ws.sheet_state = "hidden"

	for col in range(1, 25):
		charts_ws.column_dimensions[get_column_letter(col)].width = 12

	charts_ws.freeze_panes = "A3"

	CHART_START_COLS = [1, 9, 17]
	CHARTS_PER_ROW = 3
	ROW_HEIGHT_BLOCK = 16

	chart_count = 0
	region_count = 0
	chart_data_row = 1
	row_cursor = 1

	for area in areas:
		area_name = str(area.get("name", "")).strip() or "Unknown Area"
		kpis: list[dict[str, Any]] = area.get("kpis", [])

		failing: list[tuple[str, Threshold, list[tuple[datetime, float]]]] = []
		for kpi in kpis:
			kpi_name = str(kpi.get("name", "")).strip()
			if not kpi_name:
				continue
			threshold = thresholds.get(kpi_name)
			if threshold is None:
				continue

			raw_points: list[dict[str, Any]] = kpi.get("values", [])
			latest = _latest_numeric_value(raw_points)
			if latest is None:
				continue
			if _meets(threshold.operator, latest, threshold.value):
				continue

			points = _sorted_points(raw_points)
			if points:
				failing.append((kpi_name, threshold, points))

		if not failing:
			continue

		region_count += 1

		charts_ws.merge_cells(f"A{row_cursor}:X{row_cursor + 1}")
		title_cell = charts_ws[f"A{row_cursor}"]
		title_cell.value = area_name.upper()
		title_cell.font = Font(bold=True, size=16)
		title_cell.alignment = Alignment(horizontal="center", vertical="center")

		grid_start_row = row_cursor + 3

		for index, (kpi_name, threshold, points) in enumerate(failing):
			data_start_row = chart_data_row
			data_ws.cell(row=data_start_row, column=1, value="Date")
			data_ws.cell(row=data_start_row, column=2, value=kpi_name)

			for offset, (date_value, metric_value) in enumerate(points, start=1):
				data_ws.cell(row=data_start_row + offset, column=1, value=date_value)
				data_ws.cell(row=data_start_row + offset, column=2, value=metric_value)

			data_end_row = data_start_row + len(points)
			chart_data_row = data_end_row + 2

			chart = LineChart()
			chart.title = f"{kpi_name} {threshold.operator}{threshold.value}"
			chart.style = 2
			chart.height = 7.2
			chart.width = 14.5
			chart.y_axis.title = "Value"
			chart.x_axis.title = "Date"

			values_ref = Reference(data_ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
			categories_ref = Reference(
				data_ws,
				min_col=1,
				min_row=data_start_row + 1,
				max_row=data_end_row
			)
			chart.add_data(values_ref, titles_from_data=True)
			chart.set_categories(categories_ref)

			grid_row = index // CHARTS_PER_ROW
			grid_col = index % CHARTS_PER_ROW
			anchor_row = grid_start_row + (grid_row * ROW_HEIGHT_BLOCK)
			anchor_col = CHART_START_COLS[grid_col]
			anchor_cell = f"{get_column_letter(anchor_col)}{anchor_row}"

			charts_ws.add_chart(chart, anchor_cell)
			chart_count += 1

		rows_used = ((len(failing) - 1) // CHARTS_PER_ROW + 1) * ROW_HEIGHT_BLOCK
		row_cursor = grid_start_row + rows_used + 2

	output_path.parent.mkdir(parents=True, exist_ok=True)
	workbook.save(output_path)

	return {
		"output_path": str(output_path),
		"regions_with_failures": region_count,
		"charts_created": chart_count,
	}


def main() -> None:
	parser = argparse.ArgumentParser()
	parser.add_argument("--input-json", required=True)
	parser.add_argument("--output", required=True)
	parser.add_argument("--source", required=False)
	args = parser.parse_args()

	input_path = Path(args.input_json)
	output_path = Path(args.output)
	source_path = Path(args.source) if args.source else None

	payload = json.loads(input_path.read_text(encoding="utf-8-sig"))
	result = generate(payload, output_path=output_path, source_path=source_path)
	print(json.dumps(result))


if __name__ == "__main__":
	main()
